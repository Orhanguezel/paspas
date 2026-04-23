#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
import subprocess
import tempfile
import unicodedata
import uuid
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook


SQL_HEADER = """-- ============================================================================
-- PASPAS ERP V1 — XLS tabanli urun / recete / kalip / operasyon senkronizasyonu
-- Kaynak: {source_name}
-- Uretim Tarihi: {generated_date}
--
-- Notlar:
--   * Nihai urun master'i "Urun Bilgileri" sayfasindan alinir.
--   * Recete matrisi "Receteler" sayfasindan alinir.
--   * Kalip + operasyonel YM eslesmeleri "Kalıplar" sayfasindan alinir.
--   * Recete satirlarinda miktari bos olan kalemler bilincli olarak atlanir.
--   * Master'da olup matrix'te bulunmayan nihai urunler sadece urun karti olarak
--     guncellenir; recete/operasyon olusturulmaz.
-- ============================================================================
SET NAMES utf8mb4;
SET time_zone = '+00:00';
SET FOREIGN_KEY_CHECKS = 0;

"""


def deterministic_uuid(*parts: str) -> str:
    return str(uuid.uuid5(uuid.NAMESPACE_URL, "::".join(parts)))


def norm_space(value: object | None) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\xa0", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text or None


def norm_key(value: object | None) -> str | None:
    text = norm_space(value)
    if text is None:
        return None
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return text.upper()


def map_supply_type(value: object | None) -> str:
    text = norm_key(value) or ""
    if "SATIN" in text:
        return "satin_alma"
    return "uretim"


def map_unit(value: object | None) -> str:
    text = norm_key(value) or ""
    mapping = {
        "TAKIM": "takim",
        "ADET": "adet",
        "KG": "kg",
        "GRAM": "gram",
        "METRE": "metre",
    }
    return mapping.get(text, (norm_space(value) or "adet").lower())


def parse_decimal(value: object | None) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = norm_space(value)
    if not text:
        return None
    text = text.replace("TRY", "").replace("%", "").replace(".", "").replace(",", ".")
    text = text.strip()
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def sql_quote(value: object | None) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, (int, float)):
        return str(value)
    text = str(value).replace("\\", "\\\\").replace("'", "''")
    return f"'{text}'"


def sql_insert_block(table: str, columns: list[str], rows: list[tuple[object, ...]]) -> str:
    chunks: list[str] = []
    column_sql = ", ".join(f"`{col}`" for col in columns)
    for start in range(0, len(rows), 250):
        part = rows[start : start + 250]
        values_sql = ",\n".join(
            "  (" + ", ".join(sql_quote(value) for value in row) + ")" for row in part
        )
        chunks.append(f"INSERT INTO `{table}` ({column_sql}) VALUES\n{values_sql};\n")
    return "\n".join(chunks)


@dataclass
class Product:
    kod: str
    kategori: str
    tedarik_tipi: str
    urun_grubu: str | None
    birim: str
    ad: str
    aciklama: str | None
    birim_fiyat: Decimal | None
    kdv_orani: Decimal
    operasyon_tipi: str | None


@dataclass
class Mold:
    kod: str
    ad: str
    aciklama: str | None = None


@dataclass
class Operation:
    ym_kod: str
    operasyon_adi: str
    kalip_kod: str | None
    hazirlik_suresi_dk: int = 60
    cevrim_suresi_sn: Decimal = Decimal("45.00")
    montaj: int = 0


@dataclass
class RecipeDef:
    urun_kod: str
    recete_kod: str
    recete_ad: str


@dataclass
class RecipeItem:
    recete_urun_kod: str
    urun_kod: str
    miktar: Decimal
    sira: int


def quantize_amount(value: Decimal) -> Decimal:
    return value.quantize(Decimal("0.0001"))


def load_master_products(workbook) -> tuple[dict[str, dict[str, object]], dict[str, str]]:
    ws = workbook["Ürün Bilgileri"]
    master: dict[str, dict[str, object]] = {}
    name_to_code: dict[str, str] = {}

    for row in range(3, ws.max_row + 1):
        kod = norm_space(ws.cell(row, 3).value)
        ad = norm_space(ws.cell(row, 5).value)
        if not kod or not ad:
            continue
        kategori = norm_key(ws.cell(row, 1).value) or ""
        kayit = {
            "kategori_text": kategori,
            "tedarik_tipi": map_supply_type(ws.cell(row, 2).value),
            "kod": kod,
            "birim": map_unit(ws.cell(row, 4).value),
            "ad": ad,
            "aciklama": norm_space(ws.cell(row, 6).value),
            "urun_grubu": norm_space(ws.cell(row, 7).value),
            "birim_fiyat": parse_decimal(ws.cell(row, 11).value),
            "kdv_orani": parse_decimal(ws.cell(row, 12).value) or Decimal("20.00"),
            "is_final": kategori.startswith("URUN"),
        }
        master[kod] = kayit
        if kayit["is_final"]:
            key = norm_key(ad)
            if key and key not in name_to_code:
                name_to_code[key] = kod

    return master, name_to_code


def fix_final_code(raw_code: str, raw_name: str | None, final_name_to_code: dict[str, str]) -> str:
    key = norm_key(raw_name)
    mapped = final_name_to_code.get(key or "")
    if mapped:
        return mapped
    return raw_code


def build_recipe_data(workbook, final_name_to_code: dict[str, str]) -> tuple[
    dict[str, dict[str, object]],
    list[RecipeDef],
    list[RecipeItem],
    set[str],
    dict[str, str],
    list[str],
]:
    ws = workbook["Reçeteler"]
    finals: dict[str, dict[str, object]] = {}
    recipe_defs: list[RecipeDef] = []
    recipe_items: list[RecipeItem] = []
    ym_codes: set[str] = set()
    ym_names: dict[str, str] = {}
    anomalies: list[str] = []

    final_item_blocks = [
        ("ym1", 3),
        ("ym2", 7),
        ("kartela", 19),
        ("barkod", 23),
        ("koli", 27),
        ("koli_etiketi", 31),
        ("cember", 35),
        ("askilik", 39),
        ("klips", 47),
        ("poset", 51),
        ("pvc_arma", 55),
        ("hologram", 59),
    ]
    ym_raw_blocks = [
        ("granul", 11),
        ("saft", 15),
        ("krom_film", 43),
    ]
    ym_recipe_profiles: dict[str, list[tuple[int, str, tuple[tuple[str, Decimal], ...]]]] = defaultdict(list)

    seen_final_codes: set[str] = set()
    for row in range(3, ws.max_row + 1):
        raw_final_kod = norm_space(ws.cell(row, 2).value)
        final_ad = norm_space(ws.cell(row, 1).value)
        if not raw_final_kod or not final_ad:
            continue

        final_kod = fix_final_code(raw_final_kod, final_ad, final_name_to_code)
        if final_kod != raw_final_kod:
            anomalies.append(f"Recete satiri kod duzeltildi: {raw_final_kod} -> {final_kod} ({final_ad})")

        ym1_kod = norm_space(ws.cell(row, 3).value)
        ym2_kod = norm_space(ws.cell(row, 7).value)
        ym1_miktar = parse_decimal(ws.cell(row, 5).value)
        ym2_miktar = parse_decimal(ws.cell(row, 9).value)
        operasyon_tipi = "cift_tarafli" if ym2_kod else "tek_tarafli"
        finals.setdefault(
            final_kod,
            {
                "ad": final_ad,
                "operasyon_tipi": operasyon_tipi,
            },
        )

        if final_kod in seen_final_codes:
            anomalies.append(f"Tekrarlayan nihai urun recete satiri: {final_kod} ({final_ad})")
            continue
        seen_final_codes.add(final_kod)

        recete_kod = f"RCT-XLS-{final_kod}"
        recete_ad = f"{final_ad} Reçetesi"
        recipe_defs.append(RecipeDef(urun_kod=final_kod, recete_kod=recete_kod, recete_ad=recete_ad))

        sira = 1
        for block_name, col in final_item_blocks:
            urun_kod = norm_space(ws.cell(row, col).value)
            urun_ad = norm_space(ws.cell(row, col + 1).value)
            miktar = parse_decimal(ws.cell(row, col + 2).value)
            if not urun_kod:
                continue

            if block_name in {"ym1", "ym2"}:
                ym_codes.add(urun_kod)
                if urun_ad:
                    ym_names.setdefault(urun_kod, urun_ad)
            elif miktar is None:
                continue

            if miktar is None:
                anomalies.append(
                    f"Miktarsiz kalem atlandi: final={final_kod}, blok={block_name}, kod={urun_kod}"
                )
                continue

            recipe_items.append(
                RecipeItem(
                    recete_urun_kod=final_kod,
                    urun_kod=urun_kod,
                    miktar=miktar,
                    sira=sira,
                )
            )
            sira += 1

        toplam_ym_miktar = Decimal("0")
        for miktar in (ym1_miktar, ym2_miktar):
            if miktar is not None:
                toplam_ym_miktar += miktar

        if toplam_ym_miktar > 0:
            per_ym_items: list[tuple[str, Decimal]] = []
            for _block_name, col in ym_raw_blocks:
                urun_kod = norm_space(ws.cell(row, col).value)
                miktar = parse_decimal(ws.cell(row, col + 2).value)
                if not urun_kod or miktar is None or miktar <= 0:
                    continue
                per_ym_items.append((urun_kod, quantize_amount(miktar / toplam_ym_miktar)))

            profile = tuple(per_ym_items)
            for ym_kod, ym_ad in ((ym1_kod, norm_space(ws.cell(row, 4).value)), (ym2_kod, norm_space(ws.cell(row, 8).value))):
                if not ym_kod:
                    continue
                ym_codes.add(ym_kod)
                if ym_ad:
                    ym_names.setdefault(ym_kod, ym_ad)
                ym_recipe_profiles[ym_kod].append((row, final_kod, profile))

    for ym_kod, profile_rows in ym_recipe_profiles.items():
        profile_counts: dict[tuple[tuple[str, Decimal], ...], int] = defaultdict(int)
        profile_first_row: dict[tuple[tuple[str, Decimal], ...], int] = {}
        profile_examples: dict[tuple[tuple[str, Decimal], ...], str] = {}

        for row_no, final_kod, profile in profile_rows:
            profile_counts[profile] += 1
            profile_first_row.setdefault(profile, row_no)
            profile_examples.setdefault(profile, final_kod)

        selected_profile = max(
            profile_counts.keys(),
            key=lambda profile: (len(profile), profile_counts[profile], -profile_first_row[profile]),
        )

        non_empty_profiles = [profile for profile in profile_counts.keys() if len(profile) > 0]
        if len(non_empty_profiles) > 1:
            anomalies.append(
                f"YM recete profili secildi: {ym_kod} (alternatif={len(non_empty_profiles)}, referans_urun={profile_examples[selected_profile]})"
            )

        if not selected_profile:
            anomalies.append(f"YM recetesi olusturulamadi: {ym_kod} (miktarli ham girdi yok)")
            continue

        recipe_defs.append(
            RecipeDef(
                urun_kod=ym_kod,
                recete_kod=f"RCT-XLS-{ym_kod}",
                recete_ad=f"{ym_names.get(ym_kod, ym_kod)} Reçetesi",
            )
        )
        for sira, (malzeme_kod, miktar) in enumerate(selected_profile, start=1):
            recipe_items.append(
                RecipeItem(
                    recete_urun_kod=ym_kod,
                    urun_kod=malzeme_kod,
                    miktar=miktar,
                    sira=sira,
                )
            )

    return finals, recipe_defs, recipe_items, ym_codes, ym_names, anomalies


def build_molds_and_operations(workbook, final_name_to_code: dict[str, str]) -> tuple[
    list[Mold],
    list[Operation],
    list[str],
]:
    ws = workbook["Kalıplar"]
    molds: dict[str, Mold] = {}
    operations: dict[str, Operation] = {}
    anomalies: list[str] = []

    for row in range(4, ws.max_row + 1):
        raw_final_kod = norm_space(ws.cell(row, 1).value)
        final_ad = norm_space(ws.cell(row, 2).value)
        if not raw_final_kod or not final_ad:
            continue

        final_kod = fix_final_code(raw_final_kod, final_ad, final_name_to_code)
        if final_kod != raw_final_kod:
            anomalies.append(f"Kalip satiri kod duzeltildi: {raw_final_kod} -> {final_kod} ({final_ad})")

        for ym_col, ym_ad_col, kalip_kod_col, kalip_ad_col in ((4, 5, 6, 7), (8, 9, 10, 11)):
            ym_kod = norm_space(ws.cell(row, ym_col).value)
            ym_ad = norm_space(ws.cell(row, ym_ad_col).value)
            kalip_kod = norm_space(ws.cell(row, kalip_kod_col).value)
            kalip_ad = norm_space(ws.cell(row, kalip_ad_col).value)
            if not ym_kod:
                continue

            if kalip_kod and kalip_ad:
                molds[kalip_kod] = Mold(kod=kalip_kod, ad=kalip_ad)

            operations[ym_kod] = Operation(
                ym_kod=ym_kod,
                operasyon_adi=ym_ad or ym_kod,
                kalip_kod=kalip_kod,
            )

    return list(molds.values()), list(operations.values()), anomalies


def build_products(
    master: dict[str, dict[str, object]],
    finals_from_recipe: dict[str, dict[str, object]],
    ym_codes: set[str],
    ym_names: dict[str, str],
) -> tuple[list[Product], list[str]]:
    products: list[Product] = []
    anomalies: list[str] = []

    for kod, row in master.items():
        is_final = bool(row["is_final"])
        ad = str(row["ad"])
        operasyon_tipi = None

        if is_final:
            kategori = "urun"
            if kod in finals_from_recipe:
                operasyon_tipi = finals_from_recipe[kod]["operasyon_tipi"]
            elif "4 PARCA" in (norm_key(ad) or ""):
                operasyon_tipi = "tek_tarafli"
                anomalies.append(f"Master'da var ama matrix'te yok; varsayilan operasyon_tipi atandi: {kod}")
        elif kod in ym_codes:
            kategori = "yarimamul"
        else:
            kategori = "hammadde"

        urun = Product(
            kod=kod,
            kategori=kategori,
            tedarik_tipi=str(row["tedarik_tipi"]),
            urun_grubu=row["urun_grubu"],
            birim=str(row["birim"]),
            ad=ym_names.get(kod) or ad,
            aciklama=row["aciklama"],
            birim_fiyat=row["birim_fiyat"],
            kdv_orani=row["kdv_orani"],
            operasyon_tipi=operasyon_tipi,
        )
        products.append(urun)

    return products, anomalies


def render_sql(
    source_path: Path,
    products: list[Product],
    molds: list[Mold],
    operations: list[Operation],
    recipe_defs: list[RecipeDef],
    recipe_items: list[RecipeItem],
    notes: list[str],
) -> str:
    sql_parts: list[str] = [
        SQL_HEADER.format(source_name=source_path.name, generated_date=datetime.now().isoformat(timespec="seconds"))
    ]

    if notes:
        sql_parts.append("-- Veri notlari:\n")
        for note in notes[:40]:
            sql_parts.append(f"--   - {note}\n")
        if len(notes) > 40:
            sql_parts.append(f"--   - ... {len(notes) - 40} ek not gizlendi\n")
        sql_parts.append("\n")

    sql_parts.append(
        """DROP TEMPORARY TABLE IF EXISTS `tmp_xls_products`;
CREATE TEMPORARY TABLE `tmp_xls_products` (
  `kod` varchar(64) NOT NULL,
  `kategori` varchar(32) NOT NULL,
  `tedarik_tipi` varchar(32) NOT NULL,
  `urun_grubu` varchar(128) DEFAULT NULL,
  `birim` varchar(16) NOT NULL,
  `ad` varchar(255) NOT NULL,
  `aciklama` varchar(500) DEFAULT NULL,
  `birim_fiyat` decimal(12,2) DEFAULT NULL,
  `kdv_orani` decimal(5,2) NOT NULL,
  `operasyon_tipi` varchar(32) DEFAULT NULL,
  PRIMARY KEY (`kod`)
) ENGINE=Memory DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;

"""
    )
    sql_parts.append(
        sql_insert_block(
            "tmp_xls_products",
            [
                "kod",
                "kategori",
                "tedarik_tipi",
                "urun_grubu",
                "birim",
                "ad",
                "aciklama",
                "birim_fiyat",
                "kdv_orani",
                "operasyon_tipi",
            ],
            [
                (
                    p.kod,
                    p.kategori,
                    p.tedarik_tipi,
                    p.urun_grubu,
                    p.birim,
                    p.ad,
                    p.aciklama,
                    p.birim_fiyat,
                    p.kdv_orani,
                    p.operasyon_tipi,
                )
                for p in sorted(products, key=lambda item: item.kod)
            ],
        )
    )

    sql_parts.append(
        """DROP TEMPORARY TABLE IF EXISTS `tmp_xls_kaliplar`;
CREATE TEMPORARY TABLE `tmp_xls_kaliplar` (
  `kod` varchar(64) NOT NULL,
  `ad` varchar(255) NOT NULL,
  `aciklama` varchar(500) DEFAULT NULL,
  PRIMARY KEY (`kod`)
) ENGINE=Memory DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;

"""
    )
    sql_parts.append(
        sql_insert_block(
            "tmp_xls_kaliplar",
            ["kod", "ad", "aciklama"],
            [(m.kod, m.ad, m.aciklama) for m in sorted(molds, key=lambda item: item.kod)],
        )
    )

    sql_parts.append(
        """DROP TEMPORARY TABLE IF EXISTS `tmp_xls_operations`;
CREATE TEMPORARY TABLE `tmp_xls_operations` (
  `id` char(36) NOT NULL,
  `ym_kod` varchar(64) NOT NULL,
  `operasyon_adi` varchar(255) NOT NULL,
  `kalip_kod` varchar(64) DEFAULT NULL,
  `hazirlik_suresi_dk` int unsigned NOT NULL,
  `cevrim_suresi_sn` decimal(10,2) NOT NULL,
  `montaj` tinyint unsigned NOT NULL,
  PRIMARY KEY (`id`),
  UNIQUE KEY `uq_tmp_xls_operations_ym` (`ym_kod`)
) ENGINE=Memory DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;

"""
    )
    sql_parts.append(
        sql_insert_block(
            "tmp_xls_operations",
            ["id", "ym_kod", "operasyon_adi", "kalip_kod", "hazirlik_suresi_dk", "cevrim_suresi_sn", "montaj"],
            [
                (
                    deterministic_uuid("xls-op", op.ym_kod),
                    op.ym_kod,
                    op.operasyon_adi,
                    op.kalip_kod,
                    op.hazirlik_suresi_dk,
                    op.cevrim_suresi_sn,
                    op.montaj,
                )
                for op in sorted(operations, key=lambda item: item.ym_kod)
            ],
        )
    )

    sql_parts.append(
        """DROP TEMPORARY TABLE IF EXISTS `tmp_xls_recipe_defs`;
CREATE TEMPORARY TABLE `tmp_xls_recipe_defs` (
  `id` char(36) NOT NULL,
  `urun_kod` varchar(64) NOT NULL,
  `recete_kod` varchar(64) NOT NULL,
  `recete_ad` varchar(255) NOT NULL,
  PRIMARY KEY (`id`),
  UNIQUE KEY `uq_tmp_xls_recipe_defs_urun` (`urun_kod`)
) ENGINE=Memory DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;

"""
    )
    sql_parts.append(
        sql_insert_block(
            "tmp_xls_recipe_defs",
            ["id", "urun_kod", "recete_kod", "recete_ad"],
            [
                (
                    deterministic_uuid("xls-recete", recipe.urun_kod),
                    recipe.urun_kod,
                    recipe.recete_kod,
                    recipe.recete_ad,
                )
                for recipe in sorted(recipe_defs, key=lambda item: item.urun_kod)
            ],
        )
    )

    sql_parts.append(
        """DROP TEMPORARY TABLE IF EXISTS `tmp_xls_recipe_items`;
CREATE TEMPORARY TABLE `tmp_xls_recipe_items` (
  `id` char(36) NOT NULL,
  `recete_id` char(36) NOT NULL,
  `urun_kod` varchar(64) NOT NULL,
  `miktar` decimal(12,4) NOT NULL,
  `sira` int unsigned NOT NULL,
  PRIMARY KEY (`id`)
) ENGINE=Memory DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;

"""
    )
    sql_parts.append(
        sql_insert_block(
            "tmp_xls_recipe_items",
            ["id", "recete_id", "urun_kod", "miktar", "sira"],
            [
                (
                    deterministic_uuid("xls-recete-kalem", item.recete_urun_kod, str(item.sira), item.urun_kod),
                    deterministic_uuid("xls-recete", item.recete_urun_kod),
                    item.urun_kod,
                    quantize_amount(item.miktar),
                    item.sira,
                )
                for item in sorted(recipe_items, key=lambda recipe_item: (recipe_item.recete_urun_kod, recipe_item.sira))
            ],
        )
    )

    sql_parts.append(
        """
-- 1) Urun master'ini kod bazli guncelle
INSERT INTO `urunler` (
  `id`, `kategori`, `tedarik_tipi`, `urun_grubu`, `kod`, `ad`, `aciklama`, `birim`,
  `stok`, `kritik_stok`, `rezerve_stok`, `birim_fiyat`, `kdv_orani`, `operasyon_tipi`,
  `is_active`, `created_at`, `updated_at`
)
SELECT
  LOWER(CONCAT(
    SUBSTR(MD5(CONCAT('xls-product-', p.kod)), 1, 8), '-',
    SUBSTR(MD5(CONCAT('xls-product-', p.kod)), 9, 4), '-',
    SUBSTR(MD5(CONCAT('xls-product-', p.kod)), 13, 4), '-',
    SUBSTR(MD5(CONCAT('xls-product-', p.kod)), 17, 4), '-',
    SUBSTR(MD5(CONCAT('xls-product-', p.kod)), 21, 12)
  )),
  p.`kategori`,
  p.`tedarik_tipi`,
  p.`urun_grubu`,
  p.`kod`,
  p.`ad`,
  p.`aciklama`,
  p.`birim`,
  0.0000,
  0.0000,
  0.0000,
  p.`birim_fiyat`,
  p.`kdv_orani`,
  p.`operasyon_tipi`,
  1,
  NOW(),
  NOW()
FROM `tmp_xls_products` p
ON DUPLICATE KEY UPDATE
  `kategori` = VALUES(`kategori`),
  `tedarik_tipi` = VALUES(`tedarik_tipi`),
  `urun_grubu` = VALUES(`urun_grubu`),
  `ad` = VALUES(`ad`),
  `aciklama` = VALUES(`aciklama`),
  `birim` = VALUES(`birim`),
  `birim_fiyat` = VALUES(`birim_fiyat`),
  `kdv_orani` = VALUES(`kdv_orani`),
  `operasyon_tipi` = VALUES(`operasyon_tipi`),
  `is_active` = 1;

-- 2) Kaliplari kod bazli guncelle
INSERT INTO `kaliplar` (`id`, `kod`, `ad`, `aciklama`, `is_active`, `created_at`, `updated_at`)
SELECT
  LOWER(CONCAT(
    SUBSTR(MD5(CONCAT('xls-kalip-', k.kod)), 1, 8), '-',
    SUBSTR(MD5(CONCAT('xls-kalip-', k.kod)), 9, 4), '-',
    SUBSTR(MD5(CONCAT('xls-kalip-', k.kod)), 13, 4), '-',
    SUBSTR(MD5(CONCAT('xls-kalip-', k.kod)), 17, 4), '-',
    SUBSTR(MD5(CONCAT('xls-kalip-', k.kod)), 21, 12)
  )),
  k.`kod`,
  k.`ad`,
  k.`aciklama`,
  1,
  NOW(),
  NOW()
FROM `tmp_xls_kaliplar` k
ON DUPLICATE KEY UPDATE
  `ad` = VALUES(`ad`),
  `aciklama` = VALUES(`aciklama`),
  `is_active` = 1;

-- 3) XLS kapsamindaki urun/yarimamul operasyonlarini sifirla
DELETE uom
FROM `urun_operasyon_makineleri` uom
INNER JOIN `urun_operasyonlari` uo ON uo.`id` = uom.`urun_operasyon_id`
INNER JOIN `urunler` u ON u.`id` = uo.`urun_id`
INNER JOIN `tmp_xls_products` p ON p.`kod` = u.`kod`
WHERE p.`kategori` IN ('urun', 'yarimamul');

DELETE uo
FROM `urun_operasyonlari` uo
INNER JOIN `urunler` u ON u.`id` = uo.`urun_id`
INNER JOIN `tmp_xls_products` p ON p.`kod` = u.`kod`
WHERE p.`kategori` IN ('urun', 'yarimamul');

-- 4) Operasyonel YM operasyonlarini yeniden kur
INSERT INTO `urun_operasyonlari` (
  `id`, `urun_id`, `sira`, `operasyon_adi`, `kalip_id`,
  `hazirlik_suresi_dk`, `cevrim_suresi_sn`, `montaj`,
  `is_active`, `created_at`, `updated_at`
)
SELECT
  o.`id`,
  u.`id`,
  1,
  o.`operasyon_adi`,
  k.`id`,
  o.`hazirlik_suresi_dk`,
  o.`cevrim_suresi_sn`,
  o.`montaj`,
  1,
  NOW(),
  NOW()
FROM `tmp_xls_operations` o
INNER JOIN `urunler` u ON u.`kod` = o.`ym_kod`
LEFT JOIN `kaliplar` k ON k.`kod` = o.`kalip_kod`;

-- 5) XLS kapsamindaki urun recete kalemlerini ve recetelerini sifirla
DELETE rk
FROM `recete_kalemleri` rk
INNER JOIN `receteler` r ON r.`id` = rk.`recete_id`
INNER JOIN `urunler` u ON u.`id` = r.`urun_id`
INNER JOIN `tmp_xls_recipe_defs` d ON d.`urun_kod` = u.`kod`;

DELETE r
FROM `receteler` r
INNER JOIN `urunler` u ON u.`id` = r.`urun_id`
INNER JOIN `tmp_xls_recipe_defs` d ON d.`urun_kod` = u.`kod`;

-- 6) Urun recetelerini matrix'e gore yeniden olustur
INSERT INTO `receteler` (
  `id`, `kod`, `ad`, `urun_id`, `aciklama`,
  `hedef_miktar`, `is_active`, `created_at`, `updated_at`
)
SELECT
  d.`id`,
  d.`recete_kod`,
  d.`recete_ad`,
  u.`id`,
  NULL,
  1.0000,
  1,
  NOW(),
  NOW()
FROM `tmp_xls_recipe_defs` d
INNER JOIN `urunler` u ON u.`kod` = d.`urun_kod`;

INSERT INTO `recete_kalemleri` (
  `id`, `recete_id`, `urun_id`, `miktar`, `fire_orani`, `sira`, `created_at`, `updated_at`
)
SELECT
  i.`id`,
  i.`recete_id`,
  u.`id`,
  i.`miktar`,
  0.00,
  i.`sira`,
  NOW(),
  NOW()
FROM `tmp_xls_recipe_items` i
INNER JOIN `urunler` u ON u.`kod` = i.`urun_kod`;

DROP TEMPORARY TABLE IF EXISTS `tmp_xls_recipe_items`;
DROP TEMPORARY TABLE IF EXISTS `tmp_xls_recipe_defs`;
DROP TEMPORARY TABLE IF EXISTS `tmp_xls_operations`;
DROP TEMPORARY TABLE IF EXISTS `tmp_xls_kaliplar`;
DROP TEMPORARY TABLE IF EXISTS `tmp_xls_products`;

SET FOREIGN_KEY_CHECKS = 1;
"""
    )
    return "".join(sql_parts)


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate XLS-based product matrix seed SQL.")
    parser.add_argument("--source", required=True, type=Path, help="Source XLS/XLSX file")
    parser.add_argument("--output", required=True, type=Path, help="Output SQL file")
    args = parser.parse_args()

    source_for_openpyxl = args.source
    temp_dir: tempfile.TemporaryDirectory[str] | None = None
    if args.source.suffix.lower() == ".xls":
        temp_dir = tempfile.TemporaryDirectory(prefix="paspas_xls_")
        subprocess.run(
            [
                "libreoffice",
                "--headless",
                "--convert-to",
                "xlsx",
                "--outdir",
                temp_dir.name,
                str(args.source),
            ],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )
        converted = Path(temp_dir.name) / f"{args.source.stem}.xlsx"
        source_for_openpyxl = converted

    workbook = load_workbook(source_for_openpyxl, data_only=True)
    master, final_name_to_code = load_master_products(workbook)
    finals_from_recipe, recipe_defs, recipe_items, ym_codes, ym_names, recipe_notes = build_recipe_data(
        workbook, final_name_to_code
    )
    molds, operations, mold_notes = build_molds_and_operations(workbook, final_name_to_code)
    products, product_notes = build_products(master, finals_from_recipe, ym_codes, ym_names)

    notes = recipe_notes + mold_notes + product_notes
    sql = render_sql(
        source_path=args.source,
        products=products,
        molds=molds,
        operations=operations,
        recipe_defs=recipe_defs,
        recipe_items=recipe_items,
        notes=notes,
    )

    args.output.write_text(sql, encoding="utf-8")

    if temp_dir is not None:
        temp_dir.cleanup()


if __name__ == "__main__":
    main()
