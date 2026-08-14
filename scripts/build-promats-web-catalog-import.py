#!/usr/bin/env python3
"""Build an idempotent Promats web-catalog SQL import from the editorial XLSX.

This intentionally targets web_promats_* tables only. ERP product tables are
never read or written. When a Drive export is supplied, its product images are
copied into the public web asset tree with stable ASCII paths.
"""

from __future__ import annotations

import argparse
import html
import re
import shutil
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass(frozen=True)
class ProductSpec:
    column: int
    slug: str
    name: str
    drive_folder: str | None
    existing: bool = False
    fallback_slug: str | None = None


PRODUCTS = (
    ProductSpec(7, "orbital-serisi", "ORBİTAL SERİSİ", None, True),
    ProductSpec(8, "maximum-serisi", "MAXIMUM SERİSİ", None, True),
    ProductSpec(9, "star-plus-serisi", "STAR PLUS SERİSİ", None, True),
    ProductSpec(10, "icon-serisi", "ICON SERİSİ", None, True),
    ProductSpec(11, "pars-serisi", "PARS SERİSİ", None, True),
    ProductSpec(12, "basak-plus-serisi", "BAŞAK PLUS SERİSİ", None, True),
    ProductSpec(13, "profesyonel-serisi", "PROFESYONEL SERİSİ", None, True),
    ProductSpec(14, "tuna-serisi", "TUNA SERİSİ", None, True),
    ProductSpec(15, "orbital-krom-serisi", "ORBİTAL KROM SERİSİ", None, fallback_slug="orbital-serisi"),
    ProductSpec(16, "kapitone-serisi", "KAPİTONE SERİSİ", "KAPITONE SERİSİ"),
    ProductSpec(17, "premium-serisi", "PREMIUM SERİSİ", "PREMIUM SERİSİ"),
    ProductSpec(18, "extra-havuzlu-serisi", "EXTRA HAVUZLU SERİSİ", "EKSTRA HAVUZLU SERİSİ"),
    ProductSpec(19, "extra-plus-serisi", "EXTRA PLUS SERİSİ", "EKSTRA PLUS SERİSİ"),
    ProductSpec(20, "extra-havuzlu-krom-serisi", "EXTRA HAVUZLU KROM SERİSİ", "EKSTRA HAVUZLU KROM SERİSİ"),
    ProductSpec(21, "extra-plus-krom-serisi", "EXTRA PLUS KROM SERİSİ", "EKSTRA PLUS KROM SERİSİ"),
    ProductSpec(22, "star-serisi", "STAR SERİSİ", "STAR SERİSİ"),
    ProductSpec(23, "gliptone-serisi", "GLIPTONE SERİSİ", "GLIPTONE SERİSİ"),
    ProductSpec(24, "badem-serisi", "BADEM SERİSİ", "BADEM SERİSİ"),
    ProductSpec(25, "yeni-nesil-serisi", "YENİ NESİL SERİSİ", "YENİ NESİL SERİSİ"),
    ProductSpec(26, "yeni-nesil-krom-serisi", "YENİ NESİL KROM SERİSİ", "YENİ NESİL KROM SERİSİ"),
)

COLOR_FILENAMES = {
    "Siyah": "color-black.png",
    "Gri": "color-gray.png",
    "Bej": "color-beige.png",
    "Karbon": "color-carbon.png",
    "Kırmızı": "color-red.png",
    "Mavi": "color-blue.png",
    "Silver": "color-silver.png",
}

DRIVE_FILENAMES = {
    "Kapak Görseli.jpg": "cover.jpg",
    "Konsept Görseli.png": "concept.png",
    "Detay Görseli.png": "detail.png",
    "Set Görseli.png": "set.png",
    **{f"Renk {color}.png": filename for color, filename in COLOR_FILENAMES.items()},
}


def text(value: object) -> str:
    return "" if value is None else str(value).strip()


def sql(value: str | None) -> str:
    if value is None or value == "":
        return "NULL"
    return "'" + value.replace("\\", "\\\\").replace("'", "''") + "'"


def plain(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def rich(value: str) -> str:
    """Turn editorial plain text into conservative, safe HTML."""
    lines = [line.strip() for line in value.replace("\r", "").split("\n")]
    out: list[str] = []
    bullets: list[str] = []

    def flush_bullets() -> None:
        if bullets:
            out.append("<ul>" + "".join(f"<li>{html.escape(item)}</li>" for item in bullets) + "</ul>")
            bullets.clear()

    for line in lines:
        if not line:
            flush_bullets()
            continue
        if line.startswith(("•", "- ")):
            bullets.append(line.lstrip("•- ").strip())
            continue
        flush_bullets()
        if ":" in line and len(line.split(":", 1)[0]) <= 32:
            label, rest = line.split(":", 1)
            out.append(f"<p><strong>{html.escape(label.strip())}:</strong>{html.escape(rest)}</p>")
        else:
            out.append(f"<p>{html.escape(line)}</p>")
    flush_bullets()
    return "".join(out)


def hero_html(value: str) -> str:
    paragraphs = [plain(part) for part in re.split(r"\n\s*\n", value) if plain(part)]
    return "<br /><br />".join(html.escape(part) for part in paragraphs)


def asset_base(slug: str) -> str:
    return f"/userfiles/images/product/catalog-2026/{slug}"


def copy_drive_assets(spec: ProductSpec, drive_root: Path, assets_root: Path) -> dict[str, str]:
    if not spec.drive_folder:
        return {}
    source = drive_root / spec.drive_folder
    if not source.is_dir():
        raise FileNotFoundError(f"Drive ürün klasörü bulunamadı: {source}")
    target = assets_root / spec.slug
    target.mkdir(parents=True, exist_ok=True)
    copied: dict[str, str] = {}
    for drive_name, target_name in DRIVE_FILENAMES.items():
        candidate = source / drive_name
        if not candidate.exists():
            continue
        shutil.copy2(candidate, target / target_name)
        copied[drive_name] = f"{asset_base(spec.slug)}/{target_name}"
    for required in ("Kapak Görseli.jpg", "Konsept Görseli.png", "Detay Görseli.png", "Set Görseli.png"):
        if required not in copied:
            raise FileNotFoundError(f"{spec.name}: zorunlu Drive görseli eksik: {required}")
    return copied


def split_hero(name: str) -> tuple[str, str]:
    suffix = " SERİSİ"
    return (name.removesuffix(suffix), "SERİSİ")


def build_sql(workbook: Path, drive_root: Path, assets_root: Path) -> str:
    ws = load_workbook(workbook, data_only=True)["Ürün İçerik Master"]
    lines = [
        "-- Generated by scripts/build-promats-web-catalog-import.py",
        "-- Scope: public Promats web catalog only; ERP tables are intentionally untouched.",
        "SET NAMES utf8mb4;",
        "START TRANSACTION;",
    ]

    for order, spec in enumerate(PRODUCTS, 1):
        c = spec.column
        values = {row: text(ws.cell(row, c).value) for row in range(2, 27)}
        hero_title1, hero_title2 = split_hero(spec.name)
        hero_description = hero_html(values[3])
        seo_description = plain(values[3])[:300]
        technical = rich(values[18])
        if values[9] or values[10]:
            technical += f"<h3>{html.escape(values[9])}</h3>{rich(values[10])}"
        advantages = f"<h3>{html.escape(values[7])}</h3>{rich(values[8])}{rich(values[19])}"
        for title_row, body_row in ((11, 12), (13, 14), (15, 16)):
            advantages += f"<h3>{html.escape(values[title_row])}</h3>{rich(values[body_row])}"
        advantages += f"<h3>Sık Sorulan Sorular</h3>{rich(values[23])}"
        advantages += f"<h3>{html.escape(values[24])}</h3>{rich(values[25])}"
        # Cleaning is explicitly marked common in the workbook. Always use the
        # common master cell: Orbital and Orbital Krom product cells currently
        # contain misplaced FAQ copy in this row.
        cleaning = f"<h3>Temizlik ve Bakım</h3>{rich(text(ws.cell(22, 6).value))}"

        copied = copy_drive_assets(spec, drive_root, assets_root)
        if copied:
            cover = copied["Kapak Görseli.jpg"]
            concept = copied["Konsept Görseli.png"]
            detail = copied["Detay Görseli.png"]
            set_image = copied["Set Görseli.png"]
        elif spec.fallback_slug:
            # Orbital Krom has no supplied Drive folder. Keep the product live by
            # reusing the approved Orbital structural imagery until originals arrive.
            base = "/images/product"
            cover = f"{base}/orbital-real-kopya.jpg"
            concept = f"{base}/orbital-havuz.png"
            detail = f"{base}/orbital-tekli.png"
            set_image = f"{base}/orbital-tam-set.png"
        else:
            cover = concept = detail = set_image = None

        lines.extend(
            [
                f"\n-- {order:02d}. {spec.name}",
                f"SET @product_id = (SELECT id FROM web_promats_products WHERE language_id=1 AND slug={sql(spec.slug)} ORDER BY id LIMIT 1);",
                "SET @new_product_id = (SELECT COALESCE(MAX(id),0)+1 FROM web_promats_products);",
                "INSERT INTO web_promats_products (id,language_id,source_language_id,sort_order,name,slug,status,created_at)",
                f"SELECT @new_product_id,1,0,{order},{sql(spec.name)},{sql(spec.slug)},1,NOW() WHERE @product_id IS NULL;",
                f"SET @product_id = (SELECT id FROM web_promats_products WHERE language_id=1 AND slug={sql(spec.slug)} ORDER BY id LIMIT 1);",
                "UPDATE web_promats_products SET",
                f"  sort_order={order}, name={sql(spec.name)},",
                f"  s1_1_text={sql(hero_title1)}, s1_2_text={sql(hero_title2)}, s1_3_text={sql(hero_description)},",
                *( [] if spec.existing else [
                    f"  s1_4_image={sql(cover)}, s2_1_image={sql(concept)},",
                    f"  s3_1_image={sql(detail)}, s3_2_image='/images/background/section7.jpg', s4_1_image={sql(set_image)},",
                ]),
                f"  s2_2_text=NULL, s2_3_text={sql(values[5])}, s2_4_text=NULL, s2_5_text={sql(values[6])},",
                f"  seo_title={sql(spec.name + ' | Promats')}, seo_description={sql(seo_description)},",
                f"  detail_description={sql(rich(values[17]))}, detail_technical={sql(technical)},",
                f"  detail_usage={sql(cleaning)}, detail_advantages={sql(advantages)},",
                f"  detail_material={sql(rich(values[21]))}, detail_universal={sql(rich(values[20]))}, status=1",
                "WHERE id=@product_id;",
            ]
        )

        if not spec.existing:
            lines.extend(
                [
                    "DELETE FROM web_promats_product_features WHERE product_id=@product_id;",
                    "SET @feature_id = (SELECT COALESCE(MAX(id),0) FROM web_promats_product_features);",
                ]
            )
            colors = [part.strip() for part in values[4].split(",") if part.strip()]
            for sort_order, color in enumerate(colors, 1):
                if copied:
                    image_path = copied.get(f"Renk {color}.png")
                else:
                    # Temporary swatches for Orbital Krom, sourced from the supplied
                    # chrome family until its own Drive folder is delivered.
                    image_path = f"/userfiles/images/product/catalog-2026/extra-havuzlu-krom-serisi/{COLOR_FILENAMES[color]}"
                lines.append(
                    "INSERT INTO web_promats_product_features "
                    "(id,product_id,type,sort_order,image,feature,status,created_at) VALUES "
                    f"((@feature_id:=@feature_id+1),@product_id,1,{sort_order},{sql(image_path)},{sql(color)},1,NOW());"
                )
            icon_images = ("/images/icon/sinif-pvc.png", "/images/icon/tescil.png", "/images/icon/ayak-dayamali.png")
            for sort_order, (title_row, icon_image) in enumerate(zip((11, 13, 15), icon_images), 1):
                lines.append(
                    "INSERT INTO web_promats_product_features "
                    "(id,product_id,type,sort_order,image,feature,status,created_at) VALUES "
                    f"((@feature_id:=@feature_id+1),@product_id,2,{sort_order},{sql(icon_image)},{sql(values[title_row])},1,NOW());"
                )
            lines.append(
                "INSERT INTO web_promats_product_features "
                "(id,product_id,type,sort_order,image,feature,status,created_at) VALUES "
                f"((@feature_id:=@feature_id+1),@product_id,3,1,'/images/set/resim1.jpg',{sql(values[10])},1,NOW());"
            )

    lines.extend(
        [
            "\n-- Correct the legacy Star Plus gray/beige image swap while updating the catalog.",
            "UPDATE web_promats_product_features f JOIN web_promats_products p ON p.id=f.product_id",
            "SET f.image=CASE WHEN f.feature LIKE 'Gri%' THEN '/images/product/star-gri-renk-002.png' WHEN f.feature LIKE 'Bej%' THEN '/images/product/star-bej-renk003.png' ELSE f.image END",
            "WHERE p.language_id=1 AND p.slug='star-plus-serisi' AND f.type=1;",
            "COMMIT;",
            "SELECT COUNT(*) AS tr_web_product_count FROM web_promats_products WHERE language_id=1;",
        ]
    )
    return "\n".join(lines) + "\n"


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--drive-root", type=Path, required=True)
    parser.add_argument("--assets-root", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    args.output.write_text(build_sql(args.workbook, args.drive_root, args.assets_root), encoding="utf-8")
    print(f"SQL: {args.output}")
    print(f"Assets: {args.assets_root}")


if __name__ == "__main__":
    main()
