#!/usr/bin/env python3
"""
SQL dump → Excel converter
Her tablo ayrı bir sheet olarak kaydedilir.
"""

import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SQL_FILE = "/home/orhan/Documents/Projeler/paspas/db_backup_20260409_2136.sql"
EXCEL_FILE = "/home/orhan/Documents/Projeler/paspas/db_backup_20260409_2136.xlsx"

# Tables to skip (system / log tables with huge data or no business value)
SKIP_TABLES = {"admin_audit_logs", "refresh_tokens", "notifications", "storage_assets"}

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def parse_sql_values(values_str: str) -> list[list]:
    """Parse VALUES clause into a list of rows (each row = list of cell values)."""
    rows = []
    # Split on '),(' boundaries carefully
    # We use a state machine to handle quotes and parentheses
    current = []
    cell = []
    in_string = False
    escape_next = False
    depth = 0
    i = 0
    s = values_str.strip()
    # Remove outer parens wrapper if present
    if s.startswith("("):
        s = s  # keep as-is, iterate char by char

    while i < len(s):
        c = s[i]

        if escape_next:
            cell.append(c)
            escape_next = False
            i += 1
            continue

        if c == "\\" and in_string:
            cell.append(c)
            escape_next = True
            i += 1
            continue

        if c == "'" and not in_string:
            in_string = True
            i += 1
            continue

        if c == "'" and in_string:
            # Check for escaped quote ''
            if i + 1 < len(s) and s[i + 1] == "'":
                cell.append("'")
                i += 2
                continue
            in_string = False
            i += 1
            continue

        if in_string:
            cell.append(c)
            i += 1
            continue

        if c == "(":
            if depth == 0:
                # Start of a new row
                depth += 1
                i += 1
                continue
            else:
                depth += 1
                cell.append(c)
                i += 1
                continue

        if c == ")":
            depth -= 1
            if depth == 0:
                # End of row
                val = "".join(cell).strip()
                if val.upper() == "NULL":
                    current.append(None)
                else:
                    current.append(val)
                rows.append(current)
                current = []
                cell = []
                i += 1
                # skip comma and whitespace
                while i < len(s) and s[i] in (",", " ", "\n", "\r", "\t"):
                    i += 1
                continue
            else:
                cell.append(c)
                i += 1
                continue

        if c == "," and depth == 1:
            val = "".join(cell).strip()
            if val.upper() == "NULL":
                current.append(None)
            else:
                current.append(val)
            cell = []
            i += 1
            continue

        cell.append(c)
        i += 1

    return rows


def parse_create_table_columns(create_sql: str) -> list[str]:
    """Extract column names from CREATE TABLE statement."""
    lines = create_sql.split("\n")
    cols = []
    for line in lines:
        line = line.strip().rstrip(",")
        # Match column definitions (backtick-quoted name)
        m = re.match(r"^`(\w+)`\s+", line)
        if m:
            cols.append(m.group(1))
    return cols


def read_sql(path: str):
    with open(path, "r", encoding="utf-8") as f:
        return f.read()


def apply_header(ws, columns: list[str]):
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
                if len(val) > max_len:
                    max_len = len(val)
            except Exception:
                pass
        # cap at 60
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


def main():
    print(f"Reading {SQL_FILE} …")
    content = read_sql(SQL_FILE)

    # Extract CREATE TABLE blocks
    create_pattern = re.compile(
        r"CREATE TABLE `(\w+)`\s*\((.*?)\)\s*ENGINE", re.DOTALL
    )
    create_blocks = {m.group(1): m.group(2) for m in create_pattern.finditer(content)}

    # Extract INSERT INTO blocks
    insert_pattern = re.compile(
        r"INSERT INTO `(\w+)` VALUES\s*(.*?);", re.DOTALL
    )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    tables_written = 0

    for m in insert_pattern.finditer(content):
        table_name = m.group(1)
        if table_name in SKIP_TABLES:
            print(f"  SKIP  {table_name}")
            continue

        values_str = m.group(2)
        rows = parse_sql_values(values_str)

        # Get columns
        columns = []
        if table_name in create_blocks:
            columns = parse_create_table_columns(create_blocks[table_name])

        print(f"  {table_name}: {len(rows)} rows, {len(columns)} cols")

        # Sheet name max 31 chars
        sheet_name = table_name[:31]
        ws = wb.create_sheet(title=sheet_name)

        if columns:
            apply_header(ws, columns)
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
        else:
            for row_idx, row_data in enumerate(rows, 1):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

        ws.freeze_panes = "A2"
        auto_width(ws)
        tables_written += 1

    # Also write empty tables (no INSERT)
    for table_name, create_sql in create_blocks.items():
        if table_name in SKIP_TABLES:
            continue
        # Check if already written
        existing = [ws.title for ws in wb.worksheets]
        if table_name[:31] in existing:
            continue
        columns = parse_create_table_columns(create_sql)
        sheet_name = table_name[:31]
        ws = wb.create_sheet(title=sheet_name)
        if columns:
            apply_header(ws, columns)
        print(f"  {table_name}: 0 rows (empty table)")
        tables_written += 1

    print(f"\nSaving to {EXCEL_FILE} …")
    wb.save(EXCEL_FILE)
    print(f"Done. {tables_written} sheets written.")


if __name__ == "__main__":
    main()
